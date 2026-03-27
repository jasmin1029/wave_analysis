"""
批量 AIC 初至拾取 + 波形分析 + 波速计算

处理流程：
1. ZERO 文件夹：对零文件，AIC 拾取 CH1/CH2 初至，ZERO到时 = CH2初至 - CH1初至
2. TEST 文件夹：
   - CH1：直接 AIC 拾取初至
   - CH2：跳过激励脉冲后 AIC 拾取
   - TEST到时 = CH2初至 - CH1初至
3. 波速计算：v = 100 mm / (TEST到时 - ZERO到时)
4. 导出 Excel：波速 + 全部波形特征参数（仅 CH2）
"""
from __future__ import annotations

import re
import sys
import time as _time
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from matplotlib import rcParams
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from first_arrival_picking import aic_method_two_stage, aic_method, compute_aic_curve
from waveform_analysis import analyze_waveform, print_analysis_report
from process_folder_csv import (
    load_scope_csv, detect_time_unit, estimate_fs,
    preprocess_voltage, find_excitation_end, safe_stem,
    plot_csv_pick, plot_waveform_analysis_figure,
)

rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "DejaVu Sans"]
rcParams["axes.unicode_minus"] = False

ROOT = Path(__file__).resolve().parent
SAMPLE_HEIGHT_MM = 100.0


def parse_channel(filename: str) -> str:
    m = re.search(r"(CH\d)", filename, re.IGNORECASE)
    return m.group(1).upper() if m else "CH?"


def parse_timestamp(filename: str) -> str:
    m = re.search(r"tek(\d{17,20})", filename, re.IGNORECASE)
    return m.group(1) if m else filename


def find_ch_pair(ch2_path: Path, folder: Path) -> Path | None:
    """根据 CH2 文件路径，找到同一时间戳的 CH1 文件。"""
    ch2_name = ch2_path.name
    ch1_name = ch2_name.replace("CH2", "CH1").replace("ch2", "ch1").replace("_CH2", "_CH1")
    ch1_path = folder / ch1_name
    if ch1_path.exists():
        return ch1_path
    for p in folder.glob("*CH1*"):
        ts2 = parse_timestamp(ch2_name)
        ts1 = parse_timestamp(p.name)
        if ts2 and ts1 and ts2 == ts1:
            return p
    return None


def pick_onset(csv_path: Path, skip_excitation: bool = False) -> dict:
    """
    AIC 拾取初至。
    skip_excitation=True 时跳过激励脉冲（用于 TEST CH2）。
    """
    time_raw, voltage = load_scope_csv(csv_path)
    time_s, _ = detect_time_unit(time_raw)
    fs = estimate_fs(time_s)
    waveform = preprocess_voltage(voltage)

    if skip_excitation:
        excitation_end = find_excitation_end(waveform, fs)
        search_waveform = waveform[excitation_end:]
        onset_local = aic_method_two_stage(
            search_waveform,
            sta_len=max(5, int(fs * 2e-6)),
            lta_len=max(50, int(fs * 20e-6)),
            trigger_ratio=3.0,
            window_expand=max(50, int(fs * 20e-6)),
        )
        onset_idx = int(np.clip(excitation_end + onset_local, 0, len(waveform) - 1))
    else:
        excitation_end = 0
        onset_idx = aic_method_two_stage(
            waveform,
            sta_len=max(5, int(fs * 2e-6)),
            lta_len=max(50, int(fs * 20e-6)),
            trigger_ratio=5.0,
            window_expand=max(50, int(fs * 20e-6)),
        )
        onset_idx = int(np.clip(onset_idx, 0, len(waveform) - 1))

    onset_time_s = float(time_s[onset_idx])
    onset_time_us = onset_time_s * 1e6
    onset_us_from_start = float((time_s[onset_idx] - time_s[0]) * 1e6)

    return {
        "file": csv_path.name,
        "channel": parse_channel(csv_path.name),
        "fs": fs,
        "n_samples": len(waveform),
        "excitation_end": excitation_end,
        "onset_idx": onset_idx,
        "onset_time_s": onset_time_s,
        "onset_time_us": onset_time_us,
        "onset_us_from_start": onset_us_from_start,
        "time_s": time_s,
        "waveform": waveform,
    }


def build_excel(zero_arrival_us: float,
                zero_ch1_us: float,
                zero_ch2_us: float,
                test_records: list[dict],
                out_path: Path) -> None:
    """导出 Excel：波速 + 波形特征参数"""
    wb = Workbook()
    ws = wb.active
    ws.title = "波速与波形特征"

    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "序号", "文件名", "通道", "时间戳",
        "采样率\n(MHz)", "数据点数",
        "TEST CH1初至\n(μs)", "TEST CH2初至\n(μs)",
        "TEST到时\n(CH2-CH1) (μs)",
        "ZERO到时\n(CH2-CH1) (μs)",
        "传播时间\n(μs)", "波速\n(m/s)",
        "最大振幅\n(V)", "峰峰值\n(V)", "RMS振幅\n(V)",
        "主频\n(kHz)", "频谱重心\n(kHz)", "带宽\n(kHz)",
        "总能量", "上升时间\n(μs)",
        "过零率", "峰值因子", "波形因子", "峭度", "偏度",
        "频谱熵\n(bits)",
        "衰减系数\n(/s)", "品质因子Q",
        "信噪比\n(dB)",
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    data_font = Font(name="Consolas", size=9)
    data_align = Alignment(horizontal="center", vertical="center")
    alt_fill = PatternFill(start_color="F0F3F5", end_color="F0F3F5", fill_type="solid")

    for row_idx, rec in enumerate(test_records, 2):
        test_arrival_us = rec["test_arrival_us"]
        travel_us = test_arrival_us - zero_arrival_us
        velocity = (SAMPLE_HEIGHT_MM / 1000.0) / (travel_us * 1e-6) if travel_us > 0 else None

        a = rec["analysis"]
        amp = a["amplitude"]
        freq = a["frequency"]
        eng = a["energy"]
        comp = a["complexity"]
        atten = a["attenuation"]

        row_data = [
            row_idx - 1,
            rec["ch2_file"],
            "CH2",
            rec["timestamp"],
            rec["fs"] / 1e6,
            rec["n_samples"],
            rec["ch1_onset_us"],
            rec["ch2_onset_us"],
            test_arrival_us,
            zero_arrival_us,
            travel_us,
            velocity,
            amp["peak_amplitude"],
            amp["peak_to_peak"],
            amp["rms_amplitude"],
            freq["dominant_freq"] / 1e3,
            freq["centroid_freq"] / 1e3,
            freq["bandwidth"] / 1e3,
            eng["total_energy"],
            eng["rise_time_s"] * 1e6,
            comp["zero_crossing_rate"],
            comp["crest_factor"],
            comp["form_factor"],
            comp["kurtosis"],
            comp["skewness"],
            comp["spectral_entropy"],
            atten["decay_coefficient"],
            atten["quality_factor_Q"],
            a["estimated_snr_db"],
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if (row_idx - 2) % 2 == 1:
                cell.fill = alt_fill

    num_fmt_map = {
        5: "0.000", 7: "0.00", 8: "0.00", 9: "0.00", 10: "0.00",
        11: "0.00", 12: "0.0",
        13: "0.000000", 14: "0.000000", 15: "0.000000",
        16: "0.00", 17: "0.00", 18: "0.00",
        19: "0.00E+00", 20: "0.00",
        21: "0.0000", 22: "0.0000", 23: "0.0000", 24: "0.0000", 25: "0.0000",
        26: "0.00",
        27: "0.0", 28: "0.0",
        29: "0.0",
    }
    for row_idx in range(2, len(test_records) + 2):
        for col_idx, fmt in num_fmt_map.items():
            ws.cell(row=row_idx, column=col_idx).number_format = fmt

    col_widths = [5, 38, 6, 20, 8, 8,
                  12, 12, 14, 14, 10, 10,
                  12, 12, 12, 10, 12, 10,
                  12, 10,
                  8, 10, 10, 10, 10, 10,
                  10, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(out_path)


def main():
    t0 = _time.time()
    zero_dir = ROOT / "ZERO"
    test_dir = ROOT / "TEST"
    out_dir = ROOT / "batch_output"
    out_dir.mkdir(parents=True, exist_ok=True)

    # ===== 1. ZERO 对零文件：拾取 CH1 和 CH2 =====
    print("=" * 70)
    print("  [1/3] 处理 ZERO 对零文件 (CH1 + CH2)")
    print("=" * 70)

    zero_ch2_csvs = sorted(zero_dir.glob("*CH2*"))
    if not zero_ch2_csvs:
        print("错误: ZERO 文件夹中未找到 CH2 CSV 文件")
        sys.exit(1)

    zero_ch2_path = zero_ch2_csvs[0]
    zero_ch1_path = find_ch_pair(zero_ch2_path, zero_dir)
    if zero_ch1_path is None:
        print("错误: ZERO 文件夹中未找到对应的 CH1 CSV 文件")
        sys.exit(1)

    print(f"  ZERO CH1: {zero_ch1_path.name}")
    zero_ch1 = pick_onset(zero_ch1_path, skip_excitation=False)
    print(f"    AIC 初至: 索引 {zero_ch1['onset_idx']}, "
          f"绝对时间 {zero_ch1['onset_time_us']:.2f} μs, "
          f"从起始 {zero_ch1['onset_us_from_start']:.2f} μs")

    print(f"  ZERO CH2: {zero_ch2_path.name}")
    zero_ch2 = pick_onset(zero_ch2_path, skip_excitation=False)
    print(f"    AIC 初至: 索引 {zero_ch2['onset_idx']}, "
          f"绝对时间 {zero_ch2['onset_time_us']:.2f} μs, "
          f"从起始 {zero_ch2['onset_us_from_start']:.2f} μs")

    zero_arrival_us = zero_ch2["onset_time_us"] - zero_ch1["onset_time_us"]
    print(f"\n  ZERO 到时 (CH2 - CH1) = {zero_ch2['onset_time_us']:.2f} - "
          f"{zero_ch1['onset_time_us']:.2f} = {zero_arrival_us:.2f} μs")

    for zr, zp in [(zero_ch1, zero_ch1_path), (zero_ch2, zero_ch2_path)]:
        zpng = out_dir / f"ZERO_{safe_stem(zp.stem)}_pick.png"
        plot_csv_pick(zp, zr["time_s"], zr["waveform"], zr["onset_idx"], zr["fs"], zpng)

    # ===== 2. TEST 文件：CH1 + CH2 配对处理 =====
    print(f"\n{'=' * 70}")
    print("  [2/3] 批量处理 TEST 波形 (CH1直接拾取 + CH2跳过激励 + 波形分析)")
    print("=" * 70)

    test_ch2_csvs = sorted(test_dir.glob("*CH2*"))
    if not test_ch2_csvs:
        print("错误: TEST 文件夹中未找到 CH2 CSV 文件")
        sys.exit(1)

    print(f"  共 {len(test_ch2_csvs)} 个 CH2 文件待处理")
    print(f"  ZERO 到时参考 (CH2-CH1): {zero_arrival_us:.2f} μs\n")

    test_records: list[dict] = []
    for i, ch2_path in enumerate(test_ch2_csvs, 1):
        label = f"[{i:3d}/{len(test_ch2_csvs)}]"
        try:
            ch1_path = find_ch_pair(ch2_path, test_dir)
            if ch1_path is None:
                print(f"  {label} {ch2_path.name:42s}  跳过: 未找到配对 CH1 文件")
                continue

            ch1_result = pick_onset(ch1_path, skip_excitation=False)
            ch2_result = pick_onset(ch2_path, skip_excitation=True)

            analysis = analyze_waveform(
                ch2_result["waveform"], ch2_result["onset_idx"], ch2_result["fs"]
            )

            ch1_onset_us = ch1_result["onset_time_us"]
            ch2_onset_us = ch2_result["onset_time_us"]
            test_arrival_us = ch2_onset_us - ch1_onset_us

            travel_us = test_arrival_us - zero_arrival_us
            velocity = (SAMPLE_HEIGHT_MM / 1000.0) / (travel_us * 1e-6) if travel_us > 0 else None
            vel_str = f"{velocity:.0f} m/s" if velocity else "N/A"

            test_records.append({
                "ch2_file": ch2_path.name,
                "ch1_file": ch1_path.name,
                "timestamp": parse_timestamp(ch2_path.name),
                "fs": ch2_result["fs"],
                "n_samples": ch2_result["n_samples"],
                "ch1_onset_us": ch1_onset_us,
                "ch2_onset_us": ch2_onset_us,
                "test_arrival_us": test_arrival_us,
                "analysis": analysis,
                "time_s": ch2_result["time_s"],
                "waveform": ch2_result["waveform"],
                "onset_idx": ch2_result["onset_idx"],
            })

            print(f"  {label} {ch2_path.name:42s}  "
                  f"CH1={ch1_onset_us:8.2f}μs  CH2={ch2_onset_us:8.2f}μs  "
                  f"到时={test_arrival_us:8.2f}μs  传播={travel_us:8.2f}μs  {vel_str}")

        except Exception as e:
            print(f"  {label} {ch2_path.name:42s}  失败: {e}")

    # ===== 3. 导出 Excel =====
    print(f"\n{'=' * 70}")
    print("  [3/3] 导出 Excel")
    print("=" * 70)

    excel_path = out_dir / "ultrasonic_velocity_analysis.xlsx"
    build_excel(
        zero_arrival_us,
        zero_ch1["onset_time_us"],
        zero_ch2["onset_time_us"],
        test_records,
        excel_path,
    )
    print(f"  Excel 已保存: {excel_path}")

    elapsed = _time.time() - t0
    n_ok = len(test_records)
    n_total = len(test_ch2_csvs)
    print(f"\n完成: {n_ok}/{n_total} 个 CH2 文件处理成功, 耗时 {elapsed:.1f} 秒")

    if test_records:
        velocities = []
        for rec in test_records:
            dt = rec["test_arrival_us"] - zero_arrival_us
            if dt > 0:
                velocities.append((SAMPLE_HEIGHT_MM / 1000.0) / (dt * 1e-6))
        if velocities:
            v_arr = np.array(velocities)
            print(f"\n  波速统计 ({len(v_arr)} 组):")
            print(f"    均值:    {np.mean(v_arr):.0f} m/s")
            print(f"    标准差:  {np.std(v_arr):.0f} m/s")
            print(f"    中位数:  {np.median(v_arr):.0f} m/s")
            print(f"    范围:    {np.min(v_arr):.0f} ~ {np.max(v_arr):.0f} m/s")


if __name__ == "__main__":
    main()
