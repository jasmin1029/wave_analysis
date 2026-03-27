"""
读取批量分析 Excel，绘制所有波形特征参数的变化曲线。
"""
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from matplotlib import rcParams
from openpyxl import load_workbook

rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "DejaVu Sans"]
rcParams["axes.unicode_minus"] = False

ROOT = Path(__file__).resolve().parent
EXCEL_PATH = ROOT / "batch_output" / "ultrasonic_velocity_analysis.xlsx"
OUT_DIR = ROOT / "batch_output"


def load_data(path: Path) -> dict[str, list]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers_raw = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [h.replace("\n", " ") if h else f"col{i}" for i, h in enumerate(headers_raw)]

    data: dict[str, list] = {h: [] for h in headers}
    for row in ws.iter_rows(min_row=2, values_only=True):
        for h, val in zip(headers, row):
            data[h].append(val)
    wb.close()
    return data


PARAM_GROUPS = [
    {
        "title": "波速与传播时间",
        "params": [
            ("波速 (m/s)", "red"),
            ("传播时间 (μs)", "steelblue"),
        ],
    },
    {
        "title": "到时参数",
        "params": [
            ("TEST CH1初至 (μs)", "#e67e22"),
            ("TEST CH2初至 (μs)", "#2980b9"),
            ("TEST到时 (CH2-CH1) (μs)", "#27ae60"),
        ],
    },
    {
        "title": "波幅参数",
        "params": [
            ("最大振幅 (V)", "#e74c3c"),
            ("峰峰值 (V)", "#3498db"),
            ("RMS振幅 (V)", "#2ecc71"),
        ],
    },
    {
        "title": "频率参数",
        "params": [
            ("主频 (kHz)", "#e74c3c"),
            ("频谱重心 (kHz)", "#3498db"),
            ("带宽 (kHz)", "#9b59b6"),
        ],
    },
    {
        "title": "能量参数",
        "params": [
            ("总能量", "#e74c3c"),
            ("上升时间 (μs)", "#3498db"),
        ],
    },
    {
        "title": "波形复杂度 (1)",
        "params": [
            ("过零率", "#e74c3c"),
            ("峰值因子", "#3498db"),
            ("波形因子", "#2ecc71"),
        ],
    },
    {
        "title": "波形复杂度 (2)",
        "params": [
            ("峭度", "#e74c3c"),
            ("偏度", "#3498db"),
            ("频谱熵 (bits)", "#9b59b6"),
        ],
    },
    {
        "title": "衰减参数",
        "params": [
            ("衰减系数 (/s)", "#e74c3c"),
            ("品质因子Q", "#3498db"),
        ],
    },
    {
        "title": "信噪比",
        "params": [
            ("信噪比 (dB)", "#e74c3c"),
        ],
    },
]


def plot_group_figure(data: dict, x: np.ndarray, group: dict,
                      out_png: Path, fig_idx: int) -> None:
    params = group["params"]
    n = len(params)
    fig, axes = plt.subplots(n, 1, figsize=(14, 3.2 * n), constrained_layout=True)
    if n == 1:
        axes = [axes]

    fig.suptitle(f"图 {fig_idx}  {group['title']} 变化曲线",
                 fontsize=15, fontweight="bold")

    for ax, (name, color) in zip(axes, params):
        y = np.array(data[name], dtype=float)
        valid = np.isfinite(y)
        xv, yv = x[valid], y[valid]

        ax.plot(xv, yv, "o-", color=color, markersize=3, linewidth=1.0, alpha=0.85)

        if len(yv) > 5:
            window = min(7, len(yv) // 3)
            if window >= 3 and window % 2 == 0:
                window += 1
            if window >= 3:
                from scipy.signal import savgol_filter
                try:
                    smooth = savgol_filter(yv, window, 2)
                    ax.plot(xv, smooth, "-", color="black", linewidth=1.8,
                            alpha=0.6, label="趋势线")
                    ax.legend(loc="upper right", fontsize=8)
                except Exception:
                    pass

        ax.set_ylabel(name, fontsize=10)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(x[0] - 0.5, x[-1] + 0.5)

        mean_v = np.nanmean(yv)
        std_v = np.nanstd(yv)
        ax.axhline(mean_v, color="gray", ls=":", lw=1, alpha=0.5)
        ax.text(0.01, 0.95, f"均值={mean_v:.4g}  σ={std_v:.4g}",
                transform=ax.transAxes, fontsize=8, va="top",
                bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))

    axes[-1].set_xlabel("测量序号", fontsize=11)
    fig.savefig(out_png, dpi=150, bbox_inches="tight")
    plt.close(fig)


def plot_overview(data: dict, x: np.ndarray, out_png: Path) -> None:
    all_params = [
        ("波速 (m/s)", "波速"),
        ("传播时间 (μs)", "传播时间"),
        ("TEST到时 (CH2-CH1) (μs)", "TEST到时"),
        ("最大振幅 (V)", "最大振幅"),
        ("峰峰值 (V)", "峰峰值"),
        ("RMS振幅 (V)", "RMS振幅"),
        ("主频 (kHz)", "主频"),
        ("频谱重心 (kHz)", "频谱重心"),
        ("带宽 (kHz)", "带宽"),
        ("总能量", "总能量"),
        ("上升时间 (μs)", "上升时间"),
        ("过零率", "过零率"),
        ("峰值因子", "峰值因子"),
        ("波形因子", "波形因子"),
        ("峭度", "峭度"),
        ("偏度", "偏度"),
        ("频谱熵 (bits)", "频谱熵"),
        ("衰减系数 (/s)", "衰减系数"),
        ("品质因子Q", "品质因子Q"),
        ("信噪比 (dB)", "信噪比"),
    ]

    n_params = len(all_params)
    cols = 4
    rows = (n_params + cols - 1) // cols

    fig, axes = plt.subplots(rows, cols, figsize=(20, 3.0 * rows), constrained_layout=True)
    fig.suptitle("全部波形特征参数变化总览", fontsize=16, fontweight="bold")
    axes_flat = axes.flatten()

    colors = plt.cm.tab20(np.linspace(0, 1, n_params))

    for i, (key, label) in enumerate(all_params):
        ax = axes_flat[i]
        y = np.array(data[key], dtype=float)
        valid = np.isfinite(y)
        xv, yv = x[valid], y[valid]

        ax.plot(xv, yv, "-", color=colors[i], linewidth=0.8, alpha=0.8)
        ax.fill_between(xv, yv, alpha=0.1, color=colors[i])
        ax.set_title(label, fontsize=10, fontweight="bold")
        ax.tick_params(labelsize=7)
        ax.grid(True, alpha=0.2)

    for j in range(n_params, len(axes_flat)):
        axes_flat[j].set_visible(False)

    fig.savefig(out_png, dpi=150, bbox_inches="tight")
    plt.close(fig)


def main():
    print(f"读取 Excel: {EXCEL_PATH}")
    data = load_data(EXCEL_PATH)
    n = len(data["序号"])
    x = np.arange(1, n + 1)
    print(f"共 {n} 条记录\n")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    for i, group in enumerate(PARAM_GROUPS, 1):
        fname = f"param_{i:02d}_{group['title'].replace(' ', '_')}.png"
        out_png = OUT_DIR / fname
        plot_group_figure(data, x, group, out_png, i)
        print(f"  图 {i}: {group['title']} → {fname}")

    overview_png = OUT_DIR / "param_00_overview.png"
    plot_overview(data, x, overview_png)
    print(f"\n  总览图 → param_00_overview.png")
    print("\n完成!")


if __name__ == "__main__":
    main()
