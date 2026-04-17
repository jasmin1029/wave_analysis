"""
波形分析系统 - Web 界面
Flask 应用，提供单一波形分析和过程波形分析两种模式。
"""
from __future__ import annotations

import base64
import io
import re
import tempfile
import traceback
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

from matplotlib.backends.backend_pdf import PdfPages

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.font_manager as _fm
import numpy as np
from matplotlib import rcParams
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy.signal import savgol_filter

from first_arrival_picking import aic_method_two_stage, compute_aic_curve
from waveform_analysis import analyze_waveform
from process_folder_csv import (
    load_scope_csv, detect_time_unit, estimate_fs,
    preprocess_voltage, find_excitation_end,
)

# ── 中文字体注册：优先使用项目内 fonts/ 目录，再找系统字体 ──
# 将字体文件提交到仓库可确保在任何部署环境（Render/Linux）下都能正常显示中文
_FONTS_DIR = Path(__file__).parent / "fonts"
_CN_FONT_PATHS = [
    # 项目内置字体（优先，跨平台可靠）
    str(_FONTS_DIR / "wqy-microhei.ttc"),
    str(_FONTS_DIR / "wqy-zenhei.ttc"),
    str(_FONTS_DIR / "NotoSansSC-Regular.otf"),
    str(_FONTS_DIR / "NotoSansSC-Regular.ttf"),
    str(_FONTS_DIR / "simhei.ttf"),
    # Windows 系统字体
    "C:/Windows/Fonts/msyh.ttc",
    "C:/Windows/Fonts/simhei.ttf",
    "C:/Windows/Fonts/simsun.ttc",
    # Linux 系统字体（apt-get install fonts-wqy-microhei）
    "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
    "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
]
_registered_cn_fonts: list[str] = []
for _fp in _CN_FONT_PATHS:
    if Path(_fp).exists():
        try:
            _fm.fontManager.addfont(_fp)
            _name = _fm.FontProperties(fname=_fp).get_name()
            if _name not in _registered_cn_fonts:
                _registered_cn_fonts.append(_name)
        except Exception:
            pass

rcParams["font.family"] = "sans-serif"
rcParams["font.sans-serif"] = _registered_cn_fonts + ["DejaVu Sans"]
rcParams["axes.unicode_minus"] = False

SAMPLE_HEIGHT_MM = 100.0

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 500 * 1024 * 1024

_cache: dict[str, dict] = {}

PARAM_GROUPS = [
    {
        "title": "波速与传播时间",
        "params": [
            ("波速 (m/s)", "red"),
            ("传播时间 (μs)", "steelblue"),
        ],
    },
    {
        "title": "到时参数",
        "params": [
            ("TEST CH1初至 (μs)", "#e67e22"),
            ("TEST CH2初至 (μs)", "#2980b9"),
            ("TEST到时 (μs)", "#27ae60"),
        ],
    },
    {
        "title": "波幅参数",
        "params": [
            ("最大振幅 (V)", "#e74c3c"),
            ("峰峰值 (V)", "#3498db"),
            ("RMS振幅 (V)", "#2ecc71"),
        ],
    },
    {
        "title": "频率参数",
        "params": [
            ("主频 (kHz)", "#e74c3c"),
            ("频谱重心 (kHz)", "#3498db"),
            ("带宽 (kHz)", "#9b59b6"),
        ],
    },
    {
        "title": "能量参数",
        "params": [
            ("总能量", "#e74c3c"),
            ("上升时间 (μs)", "#3498db"),
        ],
    },
    {
        "title": "波形复杂度 (1)",
        "params": [
            ("过零率", "#e74c3c"),
            ("峰值因子", "#3498db"),
            ("波形因子", "#2ecc71"),
        ],
    },
    {
        "title": "波形复杂度 (2)",
        "params": [
            ("峭度", "#e74c3c"),
            ("偏度", "#3498db"),
            ("频谱熵 (bits)", "#9b59b6"),
        ],
    },
    {
        "title": "衰减参数",
        "params": [
            ("衰减系数 (/s)", "#e74c3c"),
            ("品质因子Q", "#3498db"),
        ],
    },
    {
        "title": "信噪比",
        "params": [
            ("信噪比 (dB)", "#e74c3c"),
        ],
    },
]


# ==================== 辅助函数 ====================

def _parse_channel(fn: str) -> str:
    m = re.search(r"(CH\d)", fn, re.IGNORECASE)
    return m.group(1).upper() if m else "CH?"


def _parse_timestamp(fn: str) -> str:
    m = re.search(r"tek(\d{17,20})", fn, re.IGNORECASE)
    return m.group(1) if m else fn


def _find_ch1(ch2_path: Path, folder: Path) -> Path | None:
    n = ch2_path.name
    for repl in [("CH2", "CH1"), ("ch2", "ch1"), ("_CH2", "_CH1")]:
        candidate = folder / n.replace(*repl)
        if candidate.exists():
            return candidate
    ts = _parse_timestamp(n)
    for p in folder.glob("*CH1*"):
        if _parse_timestamp(p.name) == ts:
            return p
    return None


def _pick(csv_path: Path, skip_exc: bool = False) -> dict:
    time_raw, voltage = load_scope_csv(csv_path)
    time_s, _ = detect_time_unit(time_raw)
    fs = estimate_fs(time_s)
    wf = preprocess_voltage(voltage)

    if skip_exc:
        exc_end = find_excitation_end(wf, fs)
        seg = wf[exc_end:]
        local = aic_method_two_stage(
            seg,
            sta_len=max(5, int(fs * 2e-6)),
            lta_len=max(50, int(fs * 20e-6)),
            trigger_ratio=3.0,
            window_expand=max(50, int(fs * 20e-6)),
        )
        onset = int(np.clip(exc_end + local, 0, len(wf) - 1))
    else:
        exc_end = 0
        onset = aic_method_two_stage(
            wf,
            sta_len=max(5, int(fs * 2e-6)),
            lta_len=max(50, int(fs * 20e-6)),
            trigger_ratio=5.0,
            window_expand=max(50, int(fs * 20e-6)),
        )
        onset = int(np.clip(onset, 0, len(wf) - 1))

    return {
        "file": csv_path.name,
        "fs": fs,
        "n_samples": len(wf),
        "onset_idx": onset,
        "onset_time_us": float(time_s[onset]) * 1e6,
        "time_s": time_s,
        "waveform": wf,
    }


def _fig_to_b64(fig) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode()


def _save_files(file_list, dst: Path):
    dst.mkdir(parents=True, exist_ok=True)
    for f in file_list:
        name = Path(f.filename).name
        if name.lower().endswith(".csv"):
            f.save(dst / name)


def _zero_arrival(zero_dir: Path):
    ch2s = sorted(zero_dir.glob("*CH2*"))
    if not ch2s:
        raise ValueError("ZERO 文件夹中未找到 CH2 文件")
    ch2_path = ch2s[0]
    ch1_path = _find_ch1(ch2_path, zero_dir)
    if not ch1_path:
        raise ValueError("ZERO 文件夹中未找到对应的 CH1 文件")
    ch1 = _pick(ch1_path)
    ch2 = _pick(ch2_path)
    return ch2["onset_time_us"] - ch1["onset_time_us"], ch1, ch2


def _extract_flat_params(rec, zero_arrival_us):
    """把单条 record 展开成扁平的参数字典，供绘图和 Excel 共用。"""
    ta = rec["test_arrival_us"]
    tr = ta - zero_arrival_us
    vel = (SAMPLE_HEIGHT_MM / 1000.0) / (tr * 1e-6) if tr > 0 else np.nan
    a = rec["analysis"]
    return {
        "波速 (m/s)": vel,
        "传播时间 (μs)": tr,
        "TEST CH1初至 (μs)": rec["ch1_onset_us"],
        "TEST CH2初至 (μs)": rec["ch2_onset_us"],
        "TEST到时 (μs)": ta,
        "最大振幅 (V)": a["amplitude"]["peak_amplitude"],
        "峰峰值 (V)": a["amplitude"]["peak_to_peak"],
        "RMS振幅 (V)": a["amplitude"]["rms_amplitude"],
        "主频 (kHz)": a["frequency"]["dominant_freq"] / 1e3,
        "频谱重心 (kHz)": a["frequency"]["centroid_freq"] / 1e3,
        "带宽 (kHz)": a["frequency"]["bandwidth"] / 1e3,
        "总能量": a["energy"]["total_energy"],
        "上升时间 (μs)": a["energy"]["rise_time_s"] * 1e6,
        "过零率": a["complexity"]["zero_crossing_rate"],
        "峰值因子": a["complexity"]["crest_factor"],
        "波形因子": a["complexity"]["form_factor"],
        "峭度": a["complexity"]["kurtosis"],
        "偏度": a["complexity"]["skewness"],
        "频谱熵 (bits)": a["complexity"]["spectral_entropy"],
        "衰减系数 (/s)": a["attenuation"]["decay_coefficient"],
        "品质因子Q": a["attenuation"]["quality_factor_Q"],
        "信噪比 (dB)": a["estimated_snr_db"],
    }


# ==================== 绘图函数 ====================

def _fig_aic_pick(name, time_s, wf, onset_idx):
    t_us = time_s * 1e6
    aic = compute_aic_curve(wf)
    aic_lo, aic_hi = np.nanmin(aic), np.nanmax(aic)
    span = aic_hi - aic_lo
    aic_n = (aic - aic_lo) / span if np.isfinite(span) and span > 1e-30 else np.zeros_like(aic)

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(13, 8), constrained_layout=True)
    fig.suptitle(f"AIC 初至拾取 — {name}", fontsize=14, fontweight="bold")

    ax1.plot(t_us, wf, "k-", lw=0.5, alpha=0.85, label="电压 (去直流)")
    ax1.axvline(t_us[onset_idx], color="#3498db", ls="--", lw=2,
                label=f"AIC 初至 t={t_us[onset_idx]:.2f} μs")
    ax1.set(xlabel="时间 (μs)", ylabel="电压 (V)", title="波形与初至")
    ax1.legend(loc="upper right", fontsize=9); ax1.grid(True, alpha=0.3)

    ax2.plot(t_us, aic, "b-", lw=0.7, label="AIC")
    ax2t = ax2.twinx()
    ax2t.plot(t_us, aic_n, color="gray", lw=0.4, alpha=0.6)
    ax2.axvline(t_us[onset_idx], color="#3498db", ls="--", lw=2)
    ax2.set(xlabel="时间 (μs)", ylabel="AIC 值", title="AIC 曲线与拾取位置")
    ax2.grid(True, alpha=0.3)
    return fig


def _fig_analysis(time_s, wf, ar, fs, title_sfx=""):
    onset = ar["onset_index"]
    t_us = time_s * 1e6
    freq = ar["frequency"]; eng = ar["energy"]
    atten = ar["attenuation"]; comp = ar["complexity"]; amp = ar["amplitude"]

    fig, axes = plt.subplots(3, 2, figsize=(15, 14), constrained_layout=True)
    fig.suptitle("AIC 初至后 — 波形特征分析" + title_sfx, fontsize=16, fontweight="bold")

    ax = axes[0, 0]
    ax.plot(t_us, wf, "k-", lw=0.5, alpha=0.7, label="波形")
    ax.plot(t_us[onset:], atten["envelope"], "r-", lw=1.0, alpha=0.8, label="包络线")
    ax.axvline(t_us[onset], color="#3498db", ls="--", lw=2,
               label=f"AIC 初至 ({t_us[onset]:.1f} μs)")
    ax.set(xlabel="时间 (μs)", ylabel="振幅", title="波形、包络与初至位置")
    ax.legend(fontsize=8, loc="upper right"); ax.grid(True, alpha=0.3)

    ax = axes[0, 1]
    fk = freq["freq_axis"] / 1e3; sp = freq["spectrum"]
    ax.plot(fk, sp, color="#e74c3c", lw=0.8)
    ax.axvline(freq["dominant_freq"] / 1e3, color="#3498db", ls="--", lw=1.5,
               label=f"主频 {freq['dominant_freq']/1e3:.1f} kHz")
    ax.axvline(freq["centroid_freq"] / 1e3, color="#2ecc71", ls="--", lw=1.5,
               label=f"重心 {freq['centroid_freq']/1e3:.1f} kHz")
    ax.fill_between(fk, sp, alpha=0.15, color="#e74c3c")
    ax.set(xlabel="频率 (kHz)", ylabel="幅值谱", title="频谱分析")
    ax.legend(fontsize=8); ax.grid(True, alpha=0.3)
    ax.set_xlim(0, fk[-1] * 0.6)

    ax = axes[1, 0]
    ce = eng["cumulative_energy"]
    et = np.arange(len(ce)) / fs * 1e6
    ax.plot(et, ce, color="#9b59b6", lw=1.2)
    ax.axhline(0.1, color="gray", ls=":", lw=1, alpha=0.6)
    ax.axhline(0.9, color="gray", ls=":", lw=1, alpha=0.6)
    ax.fill_between(et, ce, alpha=0.1, color="#9b59b6")
    ax.set(xlabel="初至后时间 (μs)", ylabel="归一化累积能量",
           title=f"累积能量 (上升时间: {eng['rise_time_s']*1e6:.1f} μs)")
    ax.grid(True, alpha=0.3)

    ax = axes[1, 1]
    env = atten["envelope"]; pk = int(np.argmax(env))
    ea = env[pk:]; ta = np.arange(len(ea)) / fs * 1e6
    ax.plot(ta, ea, "k-", lw=0.8, alpha=0.7, label="实际包络")
    dc = atten["decay_coefficient"]
    if dc > 0:
        tf = np.arange(len(ea)) / fs
        ax.plot(ta, env[pk] * np.exp(-dc * tf), "r--", lw=1.5,
                label=f"拟合 exp(-{dc:.0f}t)")
    ax.set(xlabel="峰值后时间 (μs)", ylabel="包络幅值",
           title=f"包络衰减 (Q = {atten['quality_factor_Q']:.1f})")
    ax.legend(fontsize=8); ax.grid(True, alpha=0.3)

    axes[2, 0].remove()
    axp = fig.add_subplot(3, 2, 5, polar=True)
    cats = ["过零率", "峰值因子", "波形因子", "脉冲因子", "峭度", "频谱熵"]
    vals = [comp["zero_crossing_rate"], comp["crest_factor"], comp["form_factor"],
            comp["impulse_factor"], abs(comp["kurtosis"]), comp["spectral_entropy"]]
    gmax = max(max(v, 1e-10) for v in vals)
    nv = [v / gmax for v in vals]
    ang = np.linspace(0, 2 * np.pi, len(cats), endpoint=False).tolist()
    nv_c = nv + [nv[0]]; ang_c = ang + [ang[0]]
    axp.plot(ang_c, nv_c, "o-", color="#e74c3c", lw=1.5, markersize=5)
    axp.fill(ang_c, nv_c, alpha=0.15, color="#e74c3c")
    axp.set_xticks(ang); axp.set_xticklabels(cats, fontsize=9)
    axp.set_title("波形复杂度 (归一化)", fontsize=11, pad=15)

    ax = axes[2, 1]; ax.axis("off")
    tbl = [
        ["参数类别", "指标", "数值"],
        ["波幅", "最大振幅", f"{amp['peak_amplitude']:.6f}"],
        ["波幅", "峰峰值", f"{amp['peak_to_peak']:.6f}"],
        ["波幅", "RMS", f"{amp['rms_amplitude']:.6f}"],
        ["频率", "主频", f"{freq['dominant_freq']/1e3:.2f} kHz"],
        ["频率", "频谱重心", f"{freq['centroid_freq']/1e3:.2f} kHz"],
        ["频率", "带宽", f"{freq['bandwidth']/1e3:.2f} kHz"],
        ["能量", "总能量", f"{eng['total_energy']:.4e}"],
        ["能量", "上升时间", f"{eng['rise_time_s']*1e6:.2f} μs"],
        ["复杂度", "过零率", f"{comp['zero_crossing_rate']:.4f}"],
        ["复杂度", "峭度", f"{comp['kurtosis']:.4f}"],
        ["复杂度", "频谱熵", f"{comp['spectral_entropy']:.2f} bits"],
        ["衰减", "衰减系数", f"{atten['decay_coefficient']:.1f} /s"],
        ["衰减", "Q", f"{atten['quality_factor_Q']:.1f}"],
        ["SNR", "估算", f"{ar['estimated_snr_db']:.1f} dB"],
    ]
    table = ax.table(cellText=tbl, loc="center", cellLoc="center")
    table.auto_set_font_size(False); table.set_fontsize(9); table.scale(1.0, 1.5)
    for j in range(3):
        table[0, j].set_facecolor("#2c3e50")
        table[0, j].set_text_props(color="white", fontweight="bold")
    for r in range(1, len(tbl)):
        if r % 2 == 0:
            for c in range(3):
                table[r, c].set_facecolor("#f0f3f5")
    ax.set_title("特征参数汇总", fontsize=11, pad=15)
    return fig


def _fig_overview(param_data_list, x):
    keys = list(param_data_list[0].keys())
    n_p = len(keys)
    cols, rows = 4, (n_p + 3) // 4

    fig, axes = plt.subplots(rows, cols, figsize=(20, 3.0 * rows), constrained_layout=True)
    fig.suptitle("全部波形特征参数变化总览", fontsize=16, fontweight="bold")
    flat = axes.flatten()
    colors = plt.cm.tab20(np.linspace(0, 1, n_p))

    for i, key in enumerate(keys):
        ax = flat[i]
        y = np.array([p[key] for p in param_data_list], dtype=float)
        valid = np.isfinite(y); xv, yv = x[valid], y[valid]
        ax.plot(xv, yv, "-", color=colors[i], lw=0.8, alpha=0.8)
        ax.fill_between(xv, yv, alpha=0.1, color=colors[i])
        ax.set_title(key, fontsize=10, fontweight="bold")
        ax.tick_params(labelsize=7); ax.grid(True, alpha=0.2)

    for j in range(n_p, len(flat)):
        flat[j].set_visible(False)
    return fig


def _fig_group(param_data_list, x, group_idx):
    grp = PARAM_GROUPS[group_idx]
    params = grp["params"]
    n = len(params)
    fig, axes = plt.subplots(n, 1, figsize=(14, 3.6 * n), constrained_layout=True)
    if n == 1:
        axes = [axes]

    fig.patch.set_facecolor("#ffffff")
    fig.suptitle(f"{grp['title']}  变化曲线", fontsize=15, fontweight="bold",
                 color="#1a2332", y=1.01)

    for ax, (name, color) in zip(axes, params):
        y = np.array([p[name] for p in param_data_list], dtype=float)
        valid = np.isfinite(y)
        xv, yv = x[valid], y[valid]

        ax.set_facecolor("#fdfdff")

        mean_v = np.nanmean(yv) if len(yv) else np.nan
        std_v  = np.nanstd(yv)  if len(yv) else np.nan

        # ±1σ 阴影带
        if len(yv) > 2 and np.isfinite(std_v):
            ax.axhspan(mean_v - std_v, mean_v + std_v,
                       color=color, alpha=0.07, zorder=0, lw=0)

        # 曲线下方填充
        ax.fill_between(xv, yv, alpha=0.10, color=color, zorder=1)

        # 主折线：空心圆标记
        ax.plot(xv, yv, "o-", color=color,
                markersize=5, linewidth=1.6, alpha=0.9,
                markerfacecolor="white", markeredgewidth=1.8,
                markeredgecolor=color, zorder=3, label="实测值")

        # 趋势线（Savitzky-Golay）
        if len(yv) > 5:
            window = min(7, len(yv) // 3)
            if window >= 3 and window % 2 == 0:
                window += 1
            if window >= 3:
                try:
                    smooth = savgol_filter(yv, window, 2)
                    ax.plot(xv, smooth, "-", color="#2c3e50",
                            linewidth=2.2, alpha=0.50, zorder=2, label="趋势线")
                except Exception:
                    pass

        # 均值虚线
        if np.isfinite(mean_v):
            ax.axhline(mean_v, color=color, ls="--", lw=1.1, alpha=0.45, zorder=1)

        # 坐标轴样式
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_color("#d0d8e4")
        ax.spines["bottom"].set_color("#d0d8e4")
        ax.tick_params(colors="#718096", labelsize=8)
        ax.set_ylabel(name, fontsize=10, color="#4a5568", labelpad=8)
        ax.grid(axis="y", color="#e2e8f0", linewidth=0.8, zorder=0)
        ax.grid(axis="x", color="#edf2f7", linewidth=0.6, zorder=0)
        ax.set_axisbelow(True)
        ax.set_xlim(x[0] - 0.5, x[-1] + 0.5)

        # 统计标注（右上角）
        if np.isfinite(mean_v):
            ax.text(0.99, 0.03,
                    f"均值  {mean_v:.4g}     σ  {std_v:.4g}",
                    transform=ax.transAxes, fontsize=8.5, va="bottom", ha="right",
                    color="#4a5568",
                    bbox=dict(boxstyle="round,pad=0.4", facecolor="white",
                              edgecolor="#d0d8e4", alpha=0.92, linewidth=0.8))

        handles, labels = ax.get_legend_handles_labels()
        if labels:
            ax.legend(handles, labels, loc="upper left", fontsize=8,
                      framealpha=0.92, edgecolor="#d0d8e4", fancybox=True)

    axes[-1].set_xlabel("测量序号", fontsize=11, color="#4a5568")
    return fig


def _build_excel_bytes(records, zero_arrival_us, param_data_list):
    wb = Workbook()
    ws = wb.active
    ws.title = "波速与波形特征"

    hf = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(*(Side(style="thin"),) * 4)

    headers = [
        "序号", "文件名", "通道", "时间戳",
        "采样率\n(MHz)", "数据点数",
        "TEST CH1初至\n(μs)", "TEST CH2初至\n(μs)",
        "TEST到时\n(CH2-CH1) (μs)", "ZERO到时\n(CH2-CH1) (μs)",
        "传播时间\n(μs)", "波速\n(m/s)",
        "最大振幅\n(V)", "峰峰值\n(V)", "RMS振幅\n(V)",
        "主频\n(kHz)", "频谱重心\n(kHz)", "带宽\n(kHz)",
        "总能量", "上升时间\n(μs)",
        "过零率", "峰值因子", "波形因子", "峭度", "偏度",
        "频谱熵\n(bits)", "衰减系数\n(/s)", "品质因子Q", "信噪比\n(dB)",
    ]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb

    df = Font(name="Consolas", size=9)
    da = Alignment(horizontal="center", vertical="center")
    alt = PatternFill(start_color="F0F3F5", end_color="F0F3F5", fill_type="solid")

    for ri, (rec, pd) in enumerate(zip(records, param_data_list), 2):
        row = [
            ri - 1, rec.get("ch2_file", ""), "CH2", rec.get("timestamp", ""),
            rec.get("fs", 0) / 1e6, rec.get("n_samples", 0),
            pd["TEST CH1初至 (μs)"], pd["TEST CH2初至 (μs)"],
            pd["TEST到时 (μs)"], zero_arrival_us,
            pd["传播时间 (μs)"], pd["波速 (m/s)"],
            pd["最大振幅 (V)"], pd["峰峰值 (V)"], pd["RMS振幅 (V)"],
            pd["主频 (kHz)"], pd["频谱重心 (kHz)"], pd["带宽 (kHz)"],
            pd["总能量"], pd["上升时间 (μs)"],
            pd["过零率"], pd["峰值因子"], pd["波形因子"],
            pd["峭度"], pd["偏度"], pd["频谱熵 (bits)"],
            pd["衰减系数 (/s)"], pd["品质因子Q"], pd["信噪比 (dB)"],
        ]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = df; c.alignment = da; c.border = tb
            if (ri - 2) % 2 == 1:
                c.fill = alt

    widths = [5,38,6,20,8,8,12,12,14,14,10,10,12,12,12,10,12,10,12,10,8,10,10,10,10,10,10,10,10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ==================== 路由 ====================

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/analyze", methods=["POST"])
def analyze():
    mode = request.form.get("mode", "single")
    zero_files = request.files.getlist("zero_files")
    test_files = request.files.getlist("test_files")

    if not zero_files or not test_files:
        return jsonify({"error": "请上传 ZERO 和 TEST 文件夹中的文件"}), 400

    try:
        with tempfile.TemporaryDirectory() as tmp:
            zd = Path(tmp) / "ZERO"
            td = Path(tmp) / "TEST"
            _save_files(zero_files, zd)
            _save_files(test_files, td)

            if mode == "single":
                return _do_single(zd, td)
            else:
                return _do_process(zd, td)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


def _fig_aic_pick_plotly(name: str, time_s, wf, onset_idx: int):
    """Return Plotly JSON for an interactive AIC first-arrival chart.
    Returns None if plotly is not installed."""
    try:
        from plotly.subplots import make_subplots
        import plotly.graph_objects as go
    except ImportError:
        return None

    t_us = time_s * 1e6
    aic = compute_aic_curve(wf)
    aic_lo, aic_hi = float(np.nanmin(aic)), float(np.nanmax(aic))
    span = aic_hi - aic_lo
    aic_n = (aic - aic_lo) / span if (np.isfinite(span) and span > 1e-30) else np.zeros_like(aic)

    onset_t = float(t_us[onset_idx])

    # Downsample for browser performance, but keep full resolution near the onset
    n = len(t_us)
    if n > 6000:
        step = max(1, n // 6000)
        base_idx = np.arange(0, n, step)
        fine_idx = np.arange(max(0, onset_idx - 40), min(n, onset_idx + 40))
        plot_idx = np.sort(np.unique(np.concatenate([base_idx, fine_idx])))
    else:
        plot_idx = np.arange(n)

    tp  = t_us[plot_idx].tolist()
    wp  = wf[plot_idx].tolist()
    ap  = aic[plot_idx].tolist()
    anp = aic_n[plot_idx].tolist()

    wf_min, wf_max   = float(np.min(wp)), float(np.max(wp))
    aic_min, aic_max = float(np.nanmin(ap)), float(np.nanmax(ap))
    font_fam = "Microsoft YaHei, SimHei, Arial, sans-serif"

    fig = make_subplots(
        rows=2, cols=1, shared_xaxes=True,
        subplot_titles=["波形与初至", "AIC 曲线与拾取位置"],
        vertical_spacing=0.10,
    )

    # Waveform
    fig.add_trace(go.Scatter(
        x=tp, y=wp, mode="lines", name="电压 (去直流)",
        line=dict(color="#2c3e50", width=0.8), opacity=0.85,
        hovertemplate="t=%{x:.3f} μs<br>V=%{y:.6f} V<extra></extra>",
    ), row=1, col=1)

    # Onset vline (waveform subplot) — represented as a 2-point Scatter
    fig.add_trace(go.Scatter(
        x=[onset_t, onset_t], y=[wf_min, wf_max],
        mode="lines", name=f"AIC 初至 {onset_t:.2f} μs",
        line=dict(color="#3498db", width=2, dash="dash"),
    ), row=1, col=1)

    # AIC curve
    fig.add_trace(go.Scatter(
        x=tp, y=ap, mode="lines", name="AIC 曲线",
        line=dict(color="#2980b9", width=0.8),
        hovertemplate="t=%{x:.3f} μs<br>AIC=%{y:.2f}<extra></extra>",
    ), row=2, col=1)

    # Normalized AIC overlay
    fig.add_trace(go.Scatter(
        x=tp, y=anp, mode="lines", name="归一化 AIC",
        line=dict(color="gray", width=0.4), opacity=0.5,
        hovertemplate="t=%{x:.3f} μs<br>norm=%{y:.4f}<extra></extra>",
    ), row=2, col=1)

    # Onset vline (AIC subplot)
    fig.add_trace(go.Scatter(
        x=[onset_t, onset_t], y=[aic_min, aic_max],
        mode="lines", name="_onset_aic",
        line=dict(color="#3498db", width=2, dash="dash"),
        showlegend=False,
    ), row=2, col=1)

    ax_style = dict(
        showgrid=True, gridcolor="rgba(0,0,0,0.08)",
        title_font=dict(family=font_fam),
        tickfont=dict(family=font_fam),
        linecolor="rgba(0,0,0,0.2)",
    )
    fig.update_layout(
        title=dict(
            text=f"AIC 初至拾取 — {name}",
            font=dict(size=14, family=font_fam, color="#1a2332"),
        ),
        height=600,
        hovermode="x unified",
        font=dict(family=font_fam, size=11),
        plot_bgcolor="white",
        paper_bgcolor="#fafafa",
        legend=dict(
            x=0.01, y=0.99, bgcolor="rgba(255,255,255,0.92)",
            bordercolor="rgba(0,0,0,0.12)", borderwidth=1,
            font=dict(family=font_fam),
        ),
        margin=dict(l=75, r=20, t=65, b=55),
    )
    fig.update_xaxes(title_text="时间 (μs)", row=2, col=1, **ax_style)
    fig.update_xaxes(showticklabels=False, row=1, col=1,
                     **{k: v for k, v in ax_style.items() if k != "title_text"})
    fig.update_yaxes(title_text="电压 (V)", row=1, col=1, **ax_style)
    fig.update_yaxes(title_text="AIC 值",   row=2, col=1, **ax_style)

    return fig.to_json()


def _do_single(zd, td):
    za, _, _ = _zero_arrival(zd)

    ch2s = sorted(td.glob("*CH2*"))
    if not ch2s:
        return jsonify({"error": "TEST 文件夹中未找到 CH2 文件"}), 400
    ch2p = ch2s[0]
    ch1p = _find_ch1(ch2p, td)
    if not ch1p:
        return jsonify({"error": "TEST 中未找到匹配的 CH1 文件"}), 400

    c1 = _pick(ch1p)
    c2 = _pick(ch2p, skip_exc=True)
    ta = c2["onset_time_us"] - c1["onset_time_us"]
    tr = ta - za
    vel = (SAMPLE_HEIGHT_MM / 1000.0) / (tr * 1e-6) if tr > 0 else None

    ar = analyze_waveform(c2["waveform"], c2["onset_idx"], c2["fs"])

    record = {
        "test_arrival_us": ta,
        "ch1_onset_us": c1["onset_time_us"],
        "ch2_onset_us": c2["onset_time_us"],
        "ch2_file": ch2p.name,
        "timestamp": _parse_timestamp(ch2p.name),
        "fs": c2["fs"],
        "n_samples": c2["n_samples"],
        "analysis": ar,
    }
    pd = _extract_flat_params(record, za)
    sid = uuid.uuid4().hex[:12]
    _cache[sid] = {
        "records": [record],
        "param_data_list": [pd],
        "zero_arrival_us": za,
        "waveform": c2["waveform"],
        "onset_idx": c2["onset_idx"],
        "fs": c2["fs"],
        "time_s": c2["time_s"],
        "filename": ch2p.name,
    }

    aic_plotly = _fig_aic_pick_plotly(ch2p.name, c2["time_s"], c2["waveform"], c2["onset_idx"])
    f2 = _fig_analysis(c2["time_s"], c2["waveform"], ar, c2["fs"])

    images: dict = {"waveform_analysis": _fig_to_b64(f2)}
    if aic_plotly is not None:
        images["aic_pick_plotly"] = aic_plotly
    else:
        f1 = _fig_aic_pick(ch2p.name, c2["time_s"], c2["waveform"], c2["onset_idx"])
        images["aic_pick"] = _fig_to_b64(f1)

    return jsonify({
        "mode": "single",
        "session_id": sid,
        "summary": {
            "file": ch2p.name,
            "velocity_ms": round(vel, 1) if vel else None,
            "travel_time_us": round(tr, 2),
            "zero_arrival_us": round(za, 2),
            "test_arrival_us": round(ta, 2),
            "ch1_onset_us": round(c1["onset_time_us"], 2),
            "ch2_onset_us": round(c2["onset_time_us"], 2),
            "peak_amplitude": round(ar["amplitude"]["peak_amplitude"], 6),
            "peak_to_peak": round(ar["amplitude"]["peak_to_peak"], 6),
            "rms_amplitude": round(ar["amplitude"]["rms_amplitude"], 6),
            "dominant_freq_khz": round(ar["frequency"]["dominant_freq"] / 1e3, 2),
            "centroid_freq_khz": round(ar["frequency"]["centroid_freq"] / 1e3, 2),
            "bandwidth_khz": round(ar["frequency"]["bandwidth"] / 1e3, 2),
            "total_energy": float(f"{ar['energy']['total_energy']:.4e}"),
            "rise_time_us": round(ar["energy"]["rise_time_s"] * 1e6, 2),
            "snr_db": round(ar["estimated_snr_db"], 1),
        },
        "images": images,
    })


def _process_pair(ch2p: Path, td: Path) -> dict | None:
    """处理单个文件对，供并行调用。返回 None 表示跳过（找不到 CH1）。"""
    ch1p = _find_ch1(ch2p, td)
    if not ch1p:
        return None
    c1 = _pick(ch1p)
    c2 = _pick(ch2p, skip_exc=True)
    ar = analyze_waveform(c2["waveform"], c2["onset_idx"], c2["fs"])
    ta = c2["onset_time_us"] - c1["onset_time_us"]
    return {
        "test_arrival_us": ta,
        "ch1_onset_us": c1["onset_time_us"],
        "ch2_onset_us": c2["onset_time_us"],
        "ch2_file": ch2p.name,
        "timestamp": _parse_timestamp(ch2p.name),
        "fs": c2["fs"],
        "n_samples": c2["n_samples"],
        "analysis": ar,
    }


def _do_process(zd, td):
    za, _, _ = _zero_arrival(zd)

    ch2s = sorted(td.glob("*CH2*"))
    if not ch2s:
        return jsonify({"error": "TEST 文件夹中未找到 CH2 文件"}), 400

    records = []
    with ThreadPoolExecutor() as pool:
        futures = {pool.submit(_process_pair, ch2p, td): ch2p for ch2p in ch2s}
        for future in as_completed(futures):
            result = future.result()
            if result is not None:
                records.append(result)

    records.sort(key=lambda r: r["ch2_file"])

    if not records:
        return jsonify({"error": "未能成功处理任何文件对"}), 400

    n = len(records)
    x = np.arange(1, n + 1)
    param_data_list = [_extract_flat_params(r, za) for r in records]

    fig = _fig_overview(param_data_list, x)

    velocities = [p["波速 (m/s)"] for p in param_data_list if np.isfinite(p["波速 (m/s)"])]
    v = np.array(velocities) if velocities else np.array([])

    sid = uuid.uuid4().hex[:12]
    _cache[sid] = {
        "records": records,
        "param_data_list": param_data_list,
        "zero_arrival_us": za,
        "x": x,
    }

    groups_info = [{"idx": i, "title": g["title"]} for i, g in enumerate(PARAM_GROUPS)]

    return jsonify({
        "mode": "process",
        "session_id": sid,
        "groups": groups_info,
        "summary": {
            "n_files": n,
            "velocity_mean": round(float(np.mean(v)), 0) if len(v) else None,
            "velocity_std": round(float(np.std(v)), 0) if len(v) else None,
            "velocity_median": round(float(np.median(v)), 0) if len(v) else None,
            "velocity_min": round(float(np.min(v)), 0) if len(v) else None,
            "velocity_max": round(float(np.max(v)), 0) if len(v) else None,
        },
        "images": {
            "overview": _fig_to_b64(fig),
        },
    })


@app.route("/api/group_plot/<sid>/<int:idx>")
def group_plot(sid, idx):
    entry = _cache.get(sid)
    if not entry:
        return jsonify({"error": "会话已过期，请重新分析"}), 404
    if idx < 0 or idx >= len(PARAM_GROUPS):
        return jsonify({"error": "无效的分组索引"}), 400

    fig = _fig_group(entry["param_data_list"], entry["x"], idx)
    return jsonify({"image": _fig_to_b64(fig)})


@app.route("/api/export_excel/<sid>")
def export_excel(sid):
    entry = _cache.get(sid)
    if not entry:
        return jsonify({"error": "会话已过期，请重新分析"}), 404

    buf = _build_excel_bytes(
        entry["records"], entry["zero_arrival_us"], entry["param_data_list"]
    )
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="ultrasonic_velocity_analysis.xlsx",
    )


def _direct_wave_slice(waveform, onset_idx, fs):
    """
    提取直达波窗口，并生成加窗版本（用于频谱分析，防止泄露）。
    策略：以主频估算 2.0 个周期为上限，同时用包络衰减截断——
    当包络降至峰值的 8% 以下时停止，取两者中较短者。
    最少 48 个采样点。
    返回 (dw_raw, dw_windowed, end_idx)
    """
    from waveform_analysis import frequency_params
    from scipy.signal import hilbert as _hilbert, windows as _wins
    seg = waveform[onset_idx:]
    fp = frequency_params(seg, fs)
    dom_f = fp["dominant_freq"]
    if dom_f > 0:
        period_win = int(2.0 * fs / dom_f)
    else:
        period_win = min(256, len(seg))
    period_win = min(period_win, len(seg))

    # 包络衰减截断：峰值后降至 8% 以下才截断
    env = np.abs(_hilbert(seg[:period_win]))
    peak_val = env.max() if env.size else 1.0
    peak_pos = int(np.argmax(env))
    below = np.where(env < 0.08 * peak_val)[0]
    after_peak = below[below > peak_pos]
    if after_peak.size > 0:
        envelope_win = int(after_peak[0])
    else:
        envelope_win = period_win

    win_samples = max(min(period_win, envelope_win), 48)
    end_idx = onset_idx + win_samples
    dw_raw = waveform[onset_idx:end_idx]
    # Tukey 窗（alpha=0.25）：仅对头尾各 12.5% 做余弦渐变，中段保持平坦
    tukey = _wins.tukey(len(dw_raw), alpha=0.25)
    dw_windowed = dw_raw * tukey
    return dw_raw, dw_windowed, end_idx


def _fig_direct_wave(filename, time_s, waveform, onset_idx, fs):
    """绘制直达波分析图（2×2 布局）"""
    dw, dw_win, end_idx = _direct_wave_slice(waveform, onset_idx, fs)
    t_full_us = time_s * 1e6
    t_dw_us = t_full_us[onset_idx:end_idx]

    # 加窗信号用于所有特征分析（抑制频谱泄露），原始信号用于波形图展示
    ar_dw = analyze_waveform(dw_win, 0, fs)
    freq = ar_dw["frequency"]
    eng = ar_dw["energy"]
    atten = ar_dw["attenuation"]
    amp = ar_dw["amplitude"]
    comp = ar_dw["complexity"]

    fig, axes = plt.subplots(2, 2, figsize=(14, 10), constrained_layout=True)
    fig.patch.set_facecolor("#ffffff")
    fig.suptitle(f"直达波特征分析 — {filename}", fontsize=14, fontweight="bold", color="#1a2332")

    # ── 左上：完整波形 + 直达波高亮窗口
    ax = axes[0, 0]
    ax.set_facecolor("#fdfdff")
    ax.plot(t_full_us, waveform, color="#888", lw=0.5, alpha=0.6, label="完整波形")
    ax.axvspan(t_full_us[onset_idx], t_full_us[end_idx - 1],
               color="#3498db", alpha=0.15, label="直达波窗口")
    ax.axvline(t_full_us[onset_idx], color="#3498db", ls="--", lw=1.5,
               label=f"初至 {t_full_us[onset_idx]:.1f} μs")
    ax.set(xlabel="时间 (μs)", ylabel="振幅", title="完整波形 & 直达波窗口")
    ax.legend(fontsize=8, loc="upper right")
    ax.grid(True, alpha=0.2)
    for sp in ["top", "right"]:
        ax.spines[sp].set_visible(False)

    # ── 右上：放大的直达波（原始）+ 加窗信号 + 包络
    ax = axes[0, 1]
    ax.set_facecolor("#fdfdff")
    ax.plot(t_dw_us, dw, color="#2c3e50", lw=1.0, alpha=0.5, label="直达波（原始）")
    ax.plot(t_dw_us, dw_win, color="#3498db", lw=1.2, label="加 Tukey 窗后")
    ax.plot(t_dw_us, atten["envelope"], color="#e74c3c", lw=1.2, alpha=0.85, label="包络")
    ax.fill_between(t_dw_us, dw_win, alpha=0.08, color="#3498db")
    ax.set(xlabel="时间 (μs)", ylabel="振幅", title="直达波放大图（Tukey 窗）")
    ax.legend(fontsize=8, loc="upper right")
    ax.grid(True, alpha=0.2)
    for sp in ["top", "right"]:
        ax.spines[sp].set_visible(False)

    # ── 左下：频谱
    ax = axes[1, 0]
    ax.set_facecolor("#fdfdff")
    fk = freq["freq_axis"] / 1e3
    sp = freq["spectrum"]
    ax.plot(fk, sp, color="#e74c3c", lw=0.9)
    ax.fill_between(fk, sp, alpha=0.15, color="#e74c3c")
    ax.axvline(freq["dominant_freq"] / 1e3, color="#3498db", ls="--", lw=1.5,
               label=f"主频 {freq['dominant_freq']/1e3:.1f} kHz")
    ax.axvline(freq["centroid_freq"] / 1e3, color="#2ecc71", ls="--", lw=1.5,
               label=f"重心 {freq['centroid_freq']/1e3:.1f} kHz")
    ax.set(xlabel="频率 (kHz)", ylabel="幅值谱", title="直达波频谱")
    ax.set_xlim(0, max(fk[-1] * 0.6, freq["dominant_freq"] / 1e3 * 3))
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.2)
    for sp_name in ["top", "right"]:
        ax.spines[sp_name].set_visible(False)

    # ── 右下：特征参数汇总表
    ax = axes[1, 1]
    ax.set_facecolor("#fdfdff")
    ax.axis("off")
    tbl_data = [
        ["参数类别", "指标", "数值"],
        ["波幅", "最大振幅", f"{amp['peak_amplitude']:.6f}"],
        ["波幅", "峰峰值", f"{amp['peak_to_peak']:.6f}"],
        ["波幅", "RMS 振幅", f"{amp['rms_amplitude']:.6f}"],
        ["频率", "主频", f"{freq['dominant_freq']/1e3:.2f} kHz"],
        ["频率", "频谱重心", f"{freq['centroid_freq']/1e3:.2f} kHz"],
        ["频率", "带宽", f"{freq['bandwidth']/1e3:.2f} kHz"],
        ["能量", "总能量", f"{eng['total_energy']:.4e}"],
        ["能量", "上升时间", f"{eng['rise_time_s']*1e6:.2f} μs"],
        ["复杂度", "过零率", f"{comp['zero_crossing_rate']:.4f}"],
        ["复杂度", "峭度", f"{comp['kurtosis']:.4f}"],
        ["复杂度", "频谱熵", f"{comp['spectral_entropy']:.2f} bits"],
        ["衰减", "衰减系数", f"{atten['decay_coefficient']:.1f} /s"],
        ["衰减", "品质因子 Q", f"{atten['quality_factor_Q']:.1f}"],
        ["SNR", "估算信噪比", f"{ar_dw['estimated_snr_db']:.1f} dB"],
    ]
    table = ax.table(cellText=tbl_data, loc="center", cellLoc="center")
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1.0, 1.5)
    for j in range(3):
        table[0, j].set_facecolor("#2c3e50")
        table[0, j].set_text_props(color="white", fontweight="bold")
    for r in range(1, len(tbl_data)):
        if r % 2 == 0:
            for c in range(3):
                table[r, c].set_facecolor("#f0f3f5")
    ax.set_title("直达波特征参数汇总", fontsize=11, pad=15)

    return fig


@app.route("/api/direct_wave/<sid>")
def direct_wave(sid):
    entry = _cache.get(sid)
    if not entry:
        return jsonify({"error": "会话已过期，请重新分析"}), 404
    waveform = entry.get("waveform")
    if waveform is None:
        return jsonify({"error": "缓存中无波形数据，请重新分析"}), 400

    onset_idx = entry["onset_idx"]
    fs = entry["fs"]
    time_s = entry["time_s"]
    filename = entry.get("filename", "")

    fig = _fig_direct_wave(filename, time_s, waveform, onset_idx, fs)
    return jsonify({"image": _fig_to_b64(fig)})


def _build_direct_wave_excel_bytes(filename, waveform, onset_idx, fs):
    """为直达波特征参数构建 Excel 文件"""
    dw, dw_win, end_idx = _direct_wave_slice(waveform, onset_idx, fs)
    ar = analyze_waveform(dw_win, 0, fs)  # 加 Tukey 窗后分析，抑制频谱泄露
    amp = ar["amplitude"]; freq = ar["frequency"]
    eng = ar["energy"]; comp = ar["complexity"]
    atten = ar["attenuation"]

    wb = Workbook()
    ws = wb.active
    ws.title = "直达波特征参数"

    hf = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(*(Side(style="thin"),) * 4)
    df_font = Font(name="Consolas", size=9)
    da = Alignment(horizontal="center", vertical="center")
    alt = PatternFill(start_color="F0F3F5", end_color="F0F3F5", fill_type="solid")

    headers = [
        "文件名", "采样率\n(MHz)", "初至采样点",
        "直达波窗口\n(采样点数)", "直达波时长\n(μs)",
        "最大振幅", "峰峰值", "RMS振幅", "平均绝对振幅",
        "主频\n(kHz)", "频谱重心\n(kHz)", "带宽\n(kHz)",
        "总能量", "上升时间\n(μs)",
        "过零率", "峰值因子", "波形因子", "脉冲因子", "峭度", "偏度",
        "频谱熵\n(bits)", "衰减系数\n(/s)", "品质因子Q", "信噪比\n(dB)",
    ]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb

    row_vals = [
        filename, fs / 1e6, onset_idx,
        end_idx - onset_idx, (end_idx - onset_idx) / fs * 1e6,
        amp["peak_amplitude"], amp["peak_to_peak"], amp["rms_amplitude"], amp["mean_abs_amplitude"],
        freq["dominant_freq"] / 1e3, freq["centroid_freq"] / 1e3, freq["bandwidth"] / 1e3,
        eng["total_energy"], eng["rise_time_s"] * 1e6,
        comp["zero_crossing_rate"], comp["crest_factor"], comp["form_factor"],
        comp["impulse_factor"], comp["kurtosis"], comp["skewness"],
        comp["spectral_entropy"], atten["decay_coefficient"], atten["quality_factor_Q"],
        ar["estimated_snr_db"],
    ]

    for ci, val in enumerate(row_vals, 1):
        c = ws.cell(row=2, column=ci, value=round(val, 6) if isinstance(val, float) else val)
        c.font = df_font; c.alignment = da; c.border = tb

    widths = [38,9,10,12,12,14,14,12,14,10,12,10,14,10,8,10,10,10,8,8,10,10,10,10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/api/export_direct_wave_excel/<sid>")
def export_direct_wave_excel(sid):
    entry = _cache.get(sid)
    if not entry:
        return jsonify({"error": "会话已过期，请重新分析"}), 404
    waveform = entry.get("waveform")
    if waveform is None:
        return jsonify({"error": "缓存中无波形数据，请重新分析"}), 400

    buf = _build_direct_wave_excel_bytes(
        entry.get("filename", ""), entry["waveform"], entry["onset_idx"], entry["fs"]
    )
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="direct_wave_analysis.xlsx",
    )


_REPORT_W = 11.69  # A4 横向宽度（英寸），所有报告页统一此宽度


def _pdf_save(pdf: PdfPages, fig) -> None:
    """统一所有报告页宽度后保存并关闭 figure。"""
    orig_w, orig_h = fig.get_size_inches()
    fig.set_size_inches(_REPORT_W, orig_h * _REPORT_W / orig_w)
    pdf.savefig(fig, bbox_inches="tight", dpi=150)
    plt.close(fig)


def _fig_cover(title: str, lines: list[tuple[str, str]], subtitle: str = "") -> object:
    """生成白色专业封面，lines = [(标签, 值), ...]。"""
    fig = plt.figure(figsize=(_REPORT_W, 8.27))
    fig.patch.set_facecolor("white")

    # ── 顶部蓝色标题栏 (上方 24%) ──
    ax_hdr = fig.add_axes([0, 0.76, 1, 0.24])
    ax_hdr.set_xlim(0, 1); ax_hdr.set_ylim(0, 1)
    ax_hdr.set_axis_off()

    # 渐变背景：用多个矩形条模拟
    n_grad = 40
    for k in range(n_grad):
        t = k / n_grad
        r = int(0x14 + t * (0x1a - 0x14))
        g = int(0x3a + t * (0x6e - 0x3a))
        b = int(0x5c + t * (0xbf - 0x5c))
        ax_hdr.add_patch(plt.Rectangle(
            (k / n_grad, 0), 1 / n_grad, 1,
            facecolor=(r / 255, g / 255, b / 255), zorder=0, lw=0
        ))

    # 左侧装饰竖条
    ax_hdr.add_patch(plt.Rectangle((0, 0), 0.006, 1, facecolor="white", alpha=0.35, zorder=1, lw=0))

    ax_hdr.text(0.5, 0.64, title,
                transform=ax_hdr.transAxes, fontsize=28, fontweight="bold",
                color="white", ha="center", va="center",
                zorder=2)
    if subtitle:
        ax_hdr.text(0.5, 0.22, subtitle,
                    transform=ax_hdr.transAxes, fontsize=10.5,
                    color="#c8dff5", ha="center", va="center",
                    zorder=2)

    # ── 蓝灰色副标题条 (标题栏下方细带) ──
    ax_sub = fig.add_axes([0, 0.715, 1, 0.045])
    ax_sub.set_axis_off()
    ax_sub.add_patch(plt.Rectangle((0, 0), 1, 1, facecolor="#eaf2fb", lw=0))
    now = datetime.now().strftime("%Y-%m-%d  %H:%M")
    ax_sub.text(0.98, 0.5, f"生成时间：{now}",
                transform=ax_sub.transAxes, fontsize=9,
                color="#4a6fa5", ha="right", va="center")
    ax_sub.text(0.02, 0.5, "超声波形分析系统  /  Ultrasonic Waveform Analysis",
                transform=ax_sub.transAxes, fontsize=9,
                color="#4a6fa5", ha="left", va="center")

    # ── 主内容区 (白色，带浅灰边框) ──
    ax = fig.add_axes([0.04, 0.04, 0.92, 0.655])
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.set_axis_off()
    # 外框
    ax.add_patch(plt.Rectangle((0, 0), 1, 1,
                               facecolor="white", edgecolor="#d0dce8",
                               linewidth=0.8, zorder=0))

    n = len(lines)
    cols = 2 if n > 6 else 1
    per_col = (n + cols - 1) // cols
    col_w = 1.0 / cols

    # 列分隔线
    if cols == 2:
        ax.axvline(0.5, color="#d8e4ef", lw=0.8, ymin=0.04, ymax=0.96)

    CARD_H = 0.88 / per_col   # 每行卡片高度

    for i, (label, value) in enumerate(lines):
        col = i // per_col
        row = i % per_col
        x0 = col * col_w
        y_top = 0.97 - row * CARD_H

        # 奇行浅色底
        if row % 2 == 0:
            ax.add_patch(plt.Rectangle(
                (x0 + 0.005, y_top - CARD_H + 0.005),
                col_w - 0.01, CARD_H - 0.008,
                facecolor="#f5f8fc", zorder=1, lw=0
            ))

        # 行底部分隔线（除最后一行）
        if row < per_col - 1:
            ax.axhline(y_top - CARD_H + 0.005, color="#dde8f0", lw=0.6,
                       xmin=x0 + 0.01, xmax=x0 + col_w - 0.01)

        # 蓝色左侧小方块装饰
        ax.add_patch(plt.Rectangle(
            (x0 + 0.012, y_top - CARD_H * 0.55),
            0.006, CARD_H * 0.34,
            facecolor="#2e86de", zorder=2, lw=0
        ))

        mid_y = y_top - CARD_H * 0.46

        ax.text(x0 + 0.028, mid_y + CARD_H * 0.08, label,
                fontsize=9, color="#5a7a9a", ha="left", va="center",
                transform=ax.transAxes, zorder=3)
        ax.text(x0 + col_w - 0.025, mid_y - CARD_H * 0.08, value,
                fontsize=11.5, color="#1a2e45", ha="right", va="center",
                transform=ax.transAxes, fontweight="bold", zorder=3)

    # ── 底部页脚 ──
    ax_foot = fig.add_axes([0.04, 0.005, 0.92, 0.03])
    ax_foot.set_axis_off()
    ax_foot.axhline(0.9, color="#2e86de", lw=1.0, alpha=0.4)
    ax_foot.text(0.5, 0.3, "本报告由超声波形分析系统自动生成，仅供参考",
                 transform=ax_foot.transAxes, fontsize=8,
                 color="#8fa8c0", ha="center", va="center")
    return fig


def _build_report_pdf(entry: dict, mode: str) -> io.BytesIO:
    """根据缓存 entry 构建完整分析报告 PDF，返回 BytesIO。"""
    buf = io.BytesIO()
    with PdfPages(buf) as pdf:
        if mode == "single":
            _report_single(pdf, entry)
        else:
            _report_process(pdf, entry)
    buf.seek(0)
    return buf


def _add_page_header(fig, page_title: str, filename: str = "", page_no: int = 0, total: int = 0) -> None:
    """
    在 figure 顶部插入页眉条。
    策略：完全不修改现有 axes 位置，将页眉元素放置在图形坐标 y > 1.0
    的区域外，由 bbox_inches='tight' 保存时自动扩展画布包含这部分内容。
    这样可以保留 constrained_layout 的原有行列间距，消除子图内部文字重叠。
    若原 suptitle 设置在 y >= 1.0（如 y=1.01），将其下移到图内避免被页眉遮盖。
    """
    from matplotlib.patches import Rectangle as _MRect

    # suptitle 位于 y>=1.0 时会落入页眉背景区域，提前下移到图内
    if fig._suptitle is not None:
        _, sy = fig._suptitle.get_position()
        if sy >= 1.0:
            fig._suptitle.set_position((0.5, 0.975))

    tf = fig.transFigure

    # 背景矩形（y=1.0 ~ 1.065）
    bg = _MRect((0, 1.0), 1.0, 0.065, transform=tf,
                facecolor="#f0f5fb", clip_on=False, zorder=5)
    fig.add_artist(bg)

    # 蓝色底边线
    fig.add_artist(plt.Line2D(
        [0, 1], [1.003, 1.003], transform=tf,
        color="#2e86de", lw=2.0, clip_on=False, zorder=6,
    ))

    # 左侧标题文字
    fig.text(0.013, 1.038, f"超声波形分析报告  |  {page_title}",
             transform=tf, fontsize=9.5, fontweight="bold",
             color="#1a3a5c", va="center", clip_on=False, zorder=7)

    # 右侧文件名 / 页码
    right_parts = []
    if filename:
        right_parts.append(filename)
    if page_no and total:
        right_parts.append(f"第 {page_no} / {total} 页")
    elif page_no:
        right_parts.append(f"第 {page_no} 页")
    if right_parts:
        fig.text(0.987, 1.038, "  ·  ".join(right_parts),
                 transform=tf, fontsize=8.5, color="#4a6fa5",
                 va="center", ha="right", clip_on=False, zorder=7)


def _report_single(pdf: PdfPages, entry: dict):
    records = entry["records"]
    rec = records[0]
    pd0 = entry["param_data_list"][0]
    za = entry["zero_arrival_us"]
    wf = entry["waveform"]
    onset = entry["onset_idx"]
    fs = entry["fs"]
    time_s = entry["time_s"]
    filename = entry.get("filename", "—")

    vel = pd0["波速 (m/s)"]
    tr = pd0["传播时间 (μs)"]
    ar = rec["analysis"]
    amp = ar["amplitude"]
    freq = ar["frequency"]
    atten = ar["attenuation"]

    vel_str = f"{vel:.0f} m/s" if np.isfinite(vel) else "N/A"
    lines = [
        ("分析模式",       "单次波形分析"),
        ("数据文件",       filename),
        ("ZERO到时 (μs)",  f"{za:.3f}"),
        ("TEST到时 (μs)",  f"{pd0['TEST到时 (μs)']:.3f}"),
        ("传播时间 (μs)",  f"{tr:.3f}"),
        ("纵波波速",       vel_str),
        ("主频",          f"{freq['dominant_freq']/1e3:.2f} kHz"),
        ("频谱重心",       f"{freq['centroid_freq']/1e3:.2f} kHz"),
        ("带宽",           f"{freq['bandwidth']/1e3:.2f} kHz"),
        ("最大振幅 (V)",   f"{amp['peak_amplitude']:.6f}"),
        ("峰峰值 (V)",     f"{amp['peak_to_peak']:.6f}"),
        ("RMS振幅 (V)",    f"{amp['rms_amplitude']:.6f}"),
        ("衰减系数 (/s)",  f"{atten['decay_coefficient']:.2f}"),
        ("品质因子 Q",     f"{atten['quality_factor_Q']:.2f}"),
        ("信噪比 (dB)",    f"{ar['estimated_snr_db']:.1f}"),
    ]
    total_pages = 4

    cover = _fig_cover("超声波形分析报告", lines, "单次波形分析模式")
    _pdf_save(pdf, cover)

    f_aic = _fig_aic_pick(filename, time_s, wf, onset)
    _add_page_header(f_aic, "AIC 初至拾取", filename, 2, total_pages)
    _pdf_save(pdf, f_aic)

    f_wa = _fig_analysis(time_s, wf, ar, fs)
    _add_page_header(f_wa, "波形特征分析", filename, 3, total_pages)
    _pdf_save(pdf, f_wa)

    try:
        f_dw = _fig_direct_wave(filename, time_s, wf, onset, fs)
        _add_page_header(f_dw, "直达波特征分析", filename, 4, total_pages)
        _pdf_save(pdf, f_dw)
    except Exception:
        pass

    pdf.infodict()["Title"] = "超声波形分析报告 — 单次模式"
    pdf.infodict()["Subject"] = filename
    pdf.infodict()["Creator"] = "波形分析系统"


def _report_process(pdf: PdfPages, entry: dict):
    records = entry["records"]
    param_data_list = entry["param_data_list"]
    za = entry["zero_arrival_us"]
    x = entry["x"]
    n = len(records)

    velocities = [p["波速 (m/s)"] for p in param_data_list if np.isfinite(p["波速 (m/s)"])]
    v = np.array(velocities) if velocities else np.array([np.nan])
    v_mean   = float(np.mean(v))
    v_std    = float(np.std(v))
    v_median = float(np.median(v))
    v_min    = float(np.min(v))
    v_max    = float(np.max(v))
    cv       = v_std / v_mean * 100 if v_mean > 0 else float("nan")

    def _pmean(k):
        vals = [p[k] for p in param_data_list if np.isfinite(p[k])]
        return np.mean(vals) if vals else float("nan")

    lines = [
        ("分析模式",       "过程波形批量分析"),
        ("有效文件对数",   f"{n} 组"),
        ("ZERO到时 (μs)",  f"{za:.3f}"),
        ("波速均值",       f"{v_mean:.0f} m/s"),
        ("波速标准差",     f"{v_std:.0f} m/s"),
        ("变异系数 CV",    f"{cv:.2f} %"),
        ("波速中位数",     f"{v_median:.0f} m/s"),
        ("波速范围",       f"{v_min:.0f} ~ {v_max:.0f} m/s"),
        ("均值主频",       f"{_pmean('主频 (kHz)'):.2f} kHz"),
        ("均值峰峰值",     f"{_pmean('峰峰值 (V)'):.5f} V"),
        ("均值品质因子Q",  f"{_pmean('品质因子Q'):.2f}"),
        ("均值信噪比",     f"{_pmean('信噪比 (dB)'):.1f} dB"),
        ("均值衰减系数",   f"{_pmean('衰减系数 (/s)'):.2f} /s"),
        ("均值上升时间",   f"{_pmean('上升时间 (μs)'):.2f} μs"),
    ]

    trend_note = ""
    if n >= 3:
        first_v = [p["波速 (m/s)"] for p in param_data_list[:max(1, n // 3)] if np.isfinite(p["波速 (m/s)"])]
        last_v  = [p["波速 (m/s)"] for p in param_data_list[2 * n // 3:] if np.isfinite(p["波速 (m/s)"])]
        if first_v and last_v:
            diff = np.mean(last_v) - np.mean(first_v)
            if abs(diff) > v_std * 0.5:
                trend_note = f"（注：序列末段相对初段波速{'升高' if diff > 0 else '降低'} {abs(diff):.0f} m/s）"

    subtitle = f"批量分析  ·  {n} 组测量  ·  波速 {v_mean:.0f} ± {v_std:.0f} m/s{trend_note}"
    cover = _fig_cover("超声波形分析报告", lines, subtitle)
    _pdf_save(pdf, cover)

    total_pages = 2 + len(PARAM_GROUPS)

    f_ov = _fig_overview(param_data_list, x)
    _add_page_header(f_ov, "全参数变化总览", f"{n} 组测量", 2, total_pages)
    _pdf_save(pdf, f_ov)

    for idx in range(len(PARAM_GROUPS)):
        try:
            fg = _fig_group(param_data_list, x, idx)
            if fg._suptitle is not None:
                fg._suptitle.set_visible(False)
            _add_page_header(fg, PARAM_GROUPS[idx]["title"], f"{n} 组测量",
                             idx + 3, total_pages)
            _pdf_save(pdf, fg)
        except Exception:
            pass

    pdf.infodict()["Title"] = "超声波形分析报告 — 批量模式"
    pdf.infodict()["Subject"] = f"{n} 组测量"
    pdf.infodict()["Creator"] = "波形分析系统"


@app.route("/api/export_report/<sid>")
def export_report(sid):
    entry = _cache.get(sid)
    if not entry:
        return jsonify({"error": "会话已过期，请重新分析"}), 404

    mode = "single" if "waveform" in entry else "process"
    try:
        buf = _build_report_pdf(entry, mode)
    except Exception:
        traceback.print_exc()
        return jsonify({"error": "报告生成失败，请查看服务器日志"}), 500

    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"波形分析报告_{now}.pdf",
    )


if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
